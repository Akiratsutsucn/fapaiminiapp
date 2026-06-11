# -*- coding: utf-8 -*-
"""生成《法拍者联盟 三城三平台 核心数据抓取与判定逻辑》Excel 报告。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 通用样式 ──────────────────────────────────────────────
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")   # 深蓝标题
HEAD_FILL = PatternFill("solid", fgColor="2E75B6")    # 蓝表头
SUB_FILL = PatternFill("solid", fgColor="DDEBF7")     # 浅蓝分组
JD_FILL = PatternFill("solid", fgColor="FCE4D6")      # 京东 橙
ALI_FILL = PatternFill("solid", fgColor="FFF2CC")     # 阿里 黄
GPAI_FILL = PatternFill("solid", fgColor="E2EFDA")    # 公拍网 绿
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
HEAD_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="微软雅黑", size=10, color="000000")
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True, color="000000")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def put_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = TITLE_FILL
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 30


def fill_rows(ws, start_row, rows, widths, platform_col=None):
    """写入数据行，自动套样式与边框。platform_col 指定平台列做底色。"""
    for i, row in enumerate(rows):
        r = start_row + i
        for j, val in enumerate(row):
            cell = ws.cell(row=r, column=j + 1, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        if platform_col is not None:
            pv = row[platform_col]
            fill = {"京东拍卖": JD_FILL, "阿里拍卖": ALI_FILL, "公拍网": GPAI_FILL}.get(pv)
            if fill:
                ws.cell(row=r, column=platform_col + 1).fill = fill
                ws.cell(row=r, column=platform_col + 1).font = BOLD_FONT
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w


# ════════════════════════════════════════════════════════════
# Sheet 1：核心数据来源与抓取逻辑
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1.核心数据来源与抓取"
headers1 = ["核心字段", "平台", "数据来源", "抓取逻辑（技术实现）", "成功率保障", "未抓到的补救措施"]
put_title(ws1, "一、三城三平台 核心数据来源与抓取逻辑（起拍价/保证金/评估价/加价幅度/面积/房源照片）", len(headers1))
for j, h in enumerate(headers1):
    ws1.cell(row=2, column=j + 1, value=h)
style_header(ws1, 2, len(headers1))

rows1 = [
 # 起拍价
 ["起拍价\n(starting_price)", "京东拍卖", "接口 getProductBasicInfo 字段 startPrice（httpx 直连，无需签名）",
  "列表+详情接口返回结构化字段，直接取 startPrice（元）", "接口字段稳定，结构化返回，成功率高",
  "①列表 displayInitialPrice 兜底 ②全文兜底挖掘 mine_starting_price（正则:起拍价/起始价/变卖价）"],
 ["起拍价", "阿里拍卖", "MTOP initData 字段 startPrice（单位:分→元）",
  "渲染页面截取 window.__ICE_SUSPENSE_LOADER__ 的 initData JSON，取 startPrice/100", "结构化字段，initData 命中即可靠",
  "①列表 API displayInitialPrice 兜底 ②全文兜底 mine_starting_price"],
 ["起拍价", "公拍网", "详情页 HTML 正文正则",
  "BeautifulSoup 取正文，4 种标签正则(起拍价/起始价/变卖价/保留价)→表格单元→列表级", "HTML 服务端渲染稳定，多重正则覆盖",
  "全文兜底挖掘 mine_starting_price 再扫一遍"],
 # 保证金
 ["保证金\n(deposit)", "京东拍卖", "接口 getProductBasicInfo 字段 ensurePrice", "结构化字段直取（元）", "接口字段稳定",
  "全文兜底挖掘 mine_deposit（正则:保证金/竞买保证金/变卖预缴款）"],
 ["保证金", "阿里拍卖", "MTOP initData 字段 foregiftPrice（分→元）", "initData JSON 取 foregiftPrice/100", "结构化字段",
  "①列表 API bail 兜底 ②全文兜底 mine_deposit"],
 ["保证金", "公拍网", "详情页 HTML 正文正则", "正则「保证金[：:]金额(万/亿/元)」+表格单元", "HTML 稳定",
  "全文兜底 mine_deposit"],
 # 评估价
 ["评估价\n(appraisal_price)", "京东拍卖", "接口字段 assessmentPrice", "结构化字段直取（元）", "接口字段稳定",
  "缺失时用 market_deal_price 兜底；全文兜底"],
 ["评估价", "阿里拍卖", "initData consultPrice / otherPrice 中 CONSULT_PRICE", "取评估价字段；缺失回退 benefit.appraisalPrice/市场价", "结构化",
  "①referencePrice 兜底 ②市场价兜底"],
 ["评估价", "公拍网", "详情页 HTML 正文正则「评估价/市场价」", "正则提取；缺失用起拍价×1.5 估算", "HTML 稳定",
  "起拍价×1.5 估算兜底"],
 # 加价幅度
 ["加价幅度\n(increment_amount)", "京东拍卖", "接口字段 priceLowerOffset（最小加价单位）", "结构化字段直取", "接口字段稳定",
  "全文兜底挖掘 mine_increment（正则:加价幅度/竞价幅度/每次加价/加价阶梯）"],
 ["加价幅度", "阿里拍卖", "MTOP initData 字段 incrementPrice（分→元）", "initData 取 incrementPrice/100", "结构化",
  "全文兜底 mine_increment（拍卖须知正文）"],
 ["加价幅度", "公拍网", "详情页 HTML 正文正则", "正则「加价幅度[：:]金额」+表格", "HTML 稳定",
  "全文兜底 mine_increment"],
 # 面积
 ["建筑面积\n(area)", "京东拍卖", "★主接口无此字段★ 渲染「标的物详情」tab 正文提取",
  "Playwright 渲染详情页→BeautifulSoup 取纯文本→3 套正则(内联型「建筑总面积:123.08平方米」/标签后8字内/表格型「建筑面积(平方米)」表头后窗口取首个小数)，负向排除「地下建筑面积」",
  "3 套正则覆盖京东两类模板；范围过滤[5,5000]㎡排除脏值",
  "成交确认书 PDF 解析；回填脚本 backfill_fields --only-missing-area 重抓"],
 ["建筑面积", "阿里拍卖", "initData auctionHouse.buildingArea", "结构化字段；缺失用列表 extra.hArea", "结构化字段",
  "全文兜底「建筑面积[：:]数值㎡」"],
 ["建筑面积", "公拍网", "详情页 HTML 正文正则", "3层:标签后直跟数值→表格单元→任意「XX平方米」+范围过滤[5,5000]", "多重正则+范围过滤",
  "全文兜底 mine_area"],
 # 房源照片
 ["房源照片\n(image_urls)", "京东拍卖", "接口 paimaiImageResultList[].imagePath", "拼接 CDN 前缀；缺失回退列表 productImage",
  "接口返回图片数组", "列表 productImage 兜底；垃圾图(广告/二维码/logo)自动隐藏不展示"],
 ["房源照片", "阿里拍卖", "initData headMedia.imageList / imageList / pictUrl", "取图片数组，升级 CDN 尺寸后缀为原图",
  "initData 图片数组", "①列表 API headerPicUrls/pictureUrl 兜底 ②CDN 尺寸降级"],
 ["房源照片", "公拍网", "详情页 16 套 CSS 选择器(.detail-img img 等)", "BeautifulSoup 多选择器提取→缩略图降级", "16 套选择器覆盖多模板",
  "缩略图降级；完全空返回[]"],
]
fill_rows(ws1, 3, rows1, [16, 11, 34, 52, 30, 42], platform_col=1)
ws1.freeze_panes = "A3"

print("Sheet1 完成")
wb.save(r"C:\Users\Administrator\Desktop\workspace\法拍者联盟小程序\法拍者联盟_核心数据与判定逻辑.xlsx")
print("已保存(部分)")

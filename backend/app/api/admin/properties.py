"""Admin property management routes — all statuses visible."""
import math
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
import io

from ...core.database import get_session
from ...core.security import get_admin_user, check_module_permission, check_write_permission
from ...core.auction_status import effective_status, effective_status_sql
from ...models.property import Property
from ...schemas import PaginatedResponse, AdminPropertyCreate, PropertyDetail

router = APIRouter()

ALLOWED_STATUS_SORT = ("starting_price", "appraisal_price", "area", "created_at")


@router.get("", response_model=PaginatedResponse)
async def list_properties_admin(
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_module_permission("properties")),
    city_id: int | None = Query(None),
    district: str | None = Query(None),
    price_min: int | None = Query(None),
    price_max: int | None = Query(None),
    area_min: float | None = Query(None),
    area_max: float | None = Query(None),
    keyword: str | None = Query(None),
    property_type: str | None = Query(None),
    auction_status: str | None = Query(None),
    auction_round: str | None = Query(None),
    sort_by: str | None = Query(None),
    sort_order: str = Query("asc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    conditions = []

    if city_id:
        conditions.append(Property.city_id == city_id)
    if district:
        conditions.append(Property.district == district)
    if price_min is not None:
        conditions.append(Property.starting_price >= price_min)
    if price_max is not None:
        conditions.append(Property.starting_price <= price_max)
    if area_min is not None:
        conditions.append(Property.area >= area_min)
    if area_max is not None:
        conditions.append(Property.area <= area_max)
    if keyword:
        conditions.append(Property.title.contains(keyword))
    if property_type:
        conditions.append(Property.property_type == property_type)
    if auction_status:
        # 按实时计算状态筛选，与小程序端口径一致
        conditions.append(effective_status_sql() == auction_status)
    if auction_round:
        conditions.append(Property.auction_round == auction_round)

    count_q = select(func.count(Property.id))
    if conditions:
        count_q = count_q.where(and_(*conditions))
    total = (await db.execute(count_q)).scalar() or 0

    order_col = getattr(Property, sort_by, Property.created_at) if sort_by in ALLOWED_STATUS_SORT else Property.created_at
    order_clause = order_col.asc() if sort_order == "asc" else order_col.desc()

    q = select(Property).order_by(order_clause).offset((page - 1) * page_size).limit(page_size)
    if conditions:
        q = q.where(and_(*conditions))
    rows = (await db.execute(q)).scalars().all()

    items = []
    for p in rows:
        cover = next((img.image_url for img in (p.images or []) if img.is_cover and not getattr(img, "hidden", 0)), None) \
            or next((img.image_url for img in (p.images or []) if not getattr(img, "hidden", 0)), None)
        items.append({
            "id": p.id, "source_url": p.source_url, "auction_platform": p.auction_platform,
            "city_id": p.city_id, "title": p.title, "province_city": p.province_city,
            "publish_date": str(p.publish_date) if p.publish_date else None,
            "source_updated_at": str(p.source_updated_at) if p.source_updated_at else None,
            "district": p.district, "sub_district": p.sub_district, "ring_road": p.ring_road,
            "address": p.address, "community_name": p.community_name,
            "lat": p.lat, "lng": p.lng,
            "property_type": p.property_type, "area": p.area, "layout": p.layout,
            "floor_info": p.floor_info, "total_floors": p.total_floors,
            "has_elevator": p.has_elevator, "orientation": p.orientation,
            "decoration": p.decoration, "build_year": p.build_year,
            "starting_price": p.starting_price, "starting_unit_price": p.starting_unit_price,
            "appraisal_price": p.appraisal_price, "court_discount_rate": p.court_discount_rate,
            "deposit": p.deposit, "increment_amount": p.increment_amount,
            "market_deal_price": p.market_deal_price, "market_deal_unit_price": p.market_deal_unit_price,
            "market_discount_rate": p.market_discount_rate,
            "listing_min_price": p.listing_min_price, "latest_deal_unit_price": p.latest_deal_unit_price,
            "latest_total_price": p.latest_total_price, "bargain_potential": p.bargain_potential,
            "beike_latest_deal_unit_price": p.beike_latest_deal_unit_price,
            "beike_latest_deal_total_price": p.beike_latest_deal_total_price,
            "beike_latest_deal_time": str(p.beike_latest_deal_time) if p.beike_latest_deal_time else None,
            "auction_round": p.auction_round,
            "auction_status": effective_status(p.auction_status, p.auction_start_time, p.auction_end_time),
            "auction_status_stored": p.auction_status,
            "auction_start_time": str(p.auction_start_time) if p.auction_start_time else None,
            "auction_end_time": str(p.auction_end_time) if p.auction_end_time else None,
            "court_name": p.court_name, "case_number": p.case_number,
            "announcement_url": p.announcement_url, "description": p.description,
            "view_count": p.view_count, "participant_count": p.participant_count,
            "loan_support": p.loan_support, "has_attachments": p.has_attachments,
            "cover_image": cover,
            "created_at": str(p.created_at), "updated_at": str(p.updated_at),
        })
    return PaginatedResponse(
        items=items, total=total, page=page, page_size=page_size,
        total_pages=max(1, math.ceil(total / page_size)),
    )


@router.get("/export")
async def export_properties(
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_module_permission("properties")),
    city_id: int | None = Query(None),
    auction_status: str | None = Query(None),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
):
    conditions = []
    if city_id:
        conditions.append(Property.city_id == city_id)
    if auction_status:
        # 按实时计算状态筛选，与列表/小程序端口径一致
        conditions.append(effective_status_sql() == auction_status)

    q = select(Property)
    if conditions:
        q = q.where(and_(*conditions))
    rows = (await db.execute(q)).scalars().all()

    # 导出的「拍卖状态」用实时计算值，与列表/小程序端口径一致。
    # 用轻量代理覆盖 auction_status，其余字段仍透传到 getattr。
    rows = [_EffectiveStatusRow(p) for p in rows]

    # 全字段中文表头映射
    FIELDS = [
        ("id", "ID"), ("title", "标的名称"), ("auction_platform", "拍卖平台"),
        ("province_city", "省市"), ("district", "区"), ("sub_district", "板块"),
        ("ring_road", "环线"), ("address", "地址"), ("community_name", "小区名"),
        ("property_type", "物业类型"), ("area", "面积(m2)"), ("layout", "户型"),
        ("floor_info", "楼层"), ("total_floors", "总楼层"), ("has_elevator", "有无电梯"),
        ("orientation", "朝向"), ("decoration", "装修情况"), ("build_year", "建筑年代"),
        ("starting_price", "起拍价(元)"), ("starting_unit_price", "起拍单价(元/m2)"),
        ("appraisal_price", "法院评估价(元)"), ("court_discount_rate", "法院折扣率"),
        ("deposit", "保证金(元)"), ("increment_amount", "加价幅度(元)"),
        ("market_deal_price", "市场成交价(元)"), ("market_deal_unit_price", "市场成交单价"),
        ("market_discount_rate", "市场折扣率"),
        ("listing_min_price", "挂牌最低价(元)"),
        ("latest_deal_unit_price", "最新成交单价"), ("latest_total_price", "最新总价(元)"),
        ("bargain_potential", "捡漏空间(元)"),
        ("beike_latest_deal_unit_price", "贝壳最新成交单价"),
        ("beike_latest_deal_total_price", "贝壳最新成交总价(元)"),
        ("beike_latest_deal_time", "贝壳最新成交时间"),
        ("auction_round", "拍卖轮次"), ("auction_status", "拍卖状态"),
        ("auction_start_time", "开拍时间"), ("auction_end_time", "结束时间"),
        ("court_name", "拍卖法院"), ("case_number", "案号"),
        ("view_count", "围观人数"), ("participant_count", "参拍人数"),
        ("loan_support", "支持贷款"), ("has_attachments", "有无附件"),
        ("source_url", "原始链接"), ("publish_date", "发布时间"),
        ("created_at", "入库时间"), ("updated_at", "更新时间"),
    ]

    if format == "xlsx":
        return _export_xlsx(rows, FIELDS)
    return _export_csv(rows, FIELDS)


class _EffectiveStatusRow:
    """导出用行代理：auction_status 返回实时计算值，其余字段透传到底层 Property。"""

    __slots__ = ("_p", "auction_status")

    def __init__(self, p):
        self._p = p
        self.auction_status = effective_status(
            p.auction_status, p.auction_start_time, p.auction_end_time
        )

    def __getattr__(self, name):
        return getattr(self._p, name)


def _export_csv(rows, FIELDS):
    output = io.StringIO()
    output.write(",".join(h for _, h in FIELDS) + "\n")
    for p in rows:
        vals = []
        for key, _ in FIELDS:
            v = getattr(p, key, "")
            if v is None:
                v = ""
            elif isinstance(v, bool):
                v = "是" if v else "否"
            v = str(v).replace('"', '""')
            vals.append(f'"{v}"')
        output.write(",".join(vals) + "\n")

    csv_content = output.getvalue()
    output.close()
    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=properties.csv"},
    )


def _export_xlsx(rows, FIELDS):
    """生成 xlsx：表头加粗+底色，列宽自适应，数字保留为数字类型，日期格式化。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date

    wb = Workbook()
    ws = wb.active
    ws.title = "房源数据"

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写表头
    for col_idx, (_, label) in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 28

    # 写数据
    body_align = Alignment(horizontal="left", vertical="center")
    for r_idx, p in enumerate(rows, start=2):
        for c_idx, (key, _) in enumerate(FIELDS, start=1):
            v = getattr(p, key, "")
            # 类型转换
            if v is None:
                cell_val = ""
            elif isinstance(v, bool):
                cell_val = "是" if v else "否"
            elif isinstance(v, (datetime, date)):
                cell_val = v
            elif isinstance(v, (int, float)):
                cell_val = v
            else:
                cell_val = str(v)
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.alignment = body_align
            cell.border = border
            # 日期格式
            if isinstance(v, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"

    # 列宽自适应（粗略）
    for col_idx, (key, label) in enumerate(FIELDS, start=1):
        col_letter = get_column_letter(col_idx)
        # 基于字段类型估算宽度
        if key in ("title", "address", "community_name", "court_name"):
            width = 36
        elif key in ("source_url", "announcement_url"):
            width = 40
        elif "price" in key or "rate" in key or "amount" in key:
            width = 14
        elif "_at" in key or "_date" in key or "_time" in key:
            width = 18
        else:
            width = 12
        ws.column_dimensions[col_letter].width = width

    # 冻结表头
    ws.freeze_panes = "A2"

    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=properties.xlsx"},
    )


@router.get("/{property_id}")
async def get_property_admin(
    property_id: int,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_module_permission("properties")),
):
    result = await db.execute(select(Property).where(Property.id == property_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="房源不存在")
    return PropertyDetail.model_validate(p)


@router.post("", response_model=PropertyDetail)
async def create_property(
    body: AdminPropertyCreate,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    p = Property(**body.model_dump(), auction_platform="人工录入", source_url=f"manual://{body.title}")
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return PropertyDetail.model_validate(p)


@router.put("/{property_id}", response_model=PropertyDetail)
async def update_property(
    property_id: int,
    body: dict,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    result = await db.execute(select(Property).where(Property.id == property_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="房源不存在")

    editable = {
        "title", "district", "sub_district", "ring_road", "address", "community_name",
        "province_city", "city_id", "lat", "lng",
        "area", "layout", "floor_info", "total_floors", "has_elevator",
        "orientation", "decoration", "build_year",
        "property_type", "starting_price", "starting_unit_price",
        "appraisal_price", "court_discount_rate", "deposit", "increment_amount",
        "market_deal_price", "market_deal_unit_price", "market_discount_rate",
        "listing_min_price", "latest_deal_unit_price", "latest_total_price", "bargain_potential",
        "beike_latest_deal_unit_price", "beike_latest_deal_total_price", "beike_latest_deal_time",
        "auction_round", "auction_status", "auction_start_time", "auction_end_time",
        "court_name", "case_number", "announcement_url",
        "description", "view_count", "participant_count",
        "loan_support", "has_attachments",
    }
    for k, v in body.items():
        if k in editable and v is not None:
            setattr(p, k, v)

    # Handle images if provided
    if "images" in body:
        from ...models.property import PropertyImage
        existing = await db.execute(select(PropertyImage).where(PropertyImage.property_id == property_id))
        for img in existing.scalars():
            await db.delete(img)
        for i, img_data in enumerate(body["images"]):
            if img_data.get("image_url"):
                db.add(PropertyImage(
                    property_id=property_id,
                    image_url=img_data["image_url"],
                    thumb_url=img_data.get("thumb_url"),
                    sort_order=img_data.get("sort_order", i),
                    is_cover=img_data.get("is_cover", i == 0),
                    hidden=img_data.get("hidden", 0),
                    hide_reason=img_data.get("hide_reason"),
                ))

    await db.commit()
    await db.refresh(p)
    return PropertyDetail.model_validate(p)


@router.post("/images/{image_id}/toggle-hidden")
async def toggle_image_hidden(
    image_id: int,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    """切换单张图片的隐藏状态（管理员恢复误判图 / 手动隐藏垃圾图）。"""
    from ...models.property import PropertyImage
    img = (await db.execute(
        select(PropertyImage).where(PropertyImage.id == image_id)
    )).scalar_one_or_none()
    if not img:
        raise HTTPException(status_code=404, detail="图片不存在")
    img.hidden = 0 if img.hidden else 1
    if img.hidden and not img.hide_reason:
        img.hide_reason = "manual"
    if not img.hidden:
        img.hide_reason = None
    await db.commit()
    return {"id": img.id, "hidden": img.hidden}


@router.delete("/{property_id}")
async def delete_property(
    property_id: int,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    result = await db.execute(select(Property).where(Property.id == property_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="房源不存在")
    await db.delete(p)
    await db.commit()
    return {"message": "删除成功"}


@router.post("/{property_id}/refresh-community")
async def refresh_community(
    property_id: int,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    """触发该房源对应小区的贝壳数据重新抓取（同步）。"""
    result = await db.execute(select(Property).where(Property.id == property_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="房源不存在")
    if not p.community_name:
        raise HTTPException(status_code=400, detail="该房源未识别小区名称")

    # 异步调爬虫
    try:
        import sys
        from pathlib import Path
        crawler_root = Path(__file__).resolve().parents[3] / "crawler"
        if str(crawler_root.parent) not in sys.path:
            sys.path.insert(0, str(crawler_root.parent))
        from crawler.community_scraper import crawl_for_property
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"爬虫模块加载失败: {e}")

    try:
        c = await crawl_for_property(property_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"抓取失败: {e}")

    if not c:
        return {"message": "抓取未命中匹配小区，请检查小区名或在贝壳中是否存在", "matched": False}

    return {
        "message": "抓取完成",
        "matched": True,
        "community_id": c.id,
        "community_name": c.name,
        "avg_price": c.avg_price,
        "last_crawled_at": str(c.last_crawled_at) if c.last_crawled_at else None,
    }


@router.get("/{property_id}/community")
async def get_property_community(
    property_id: int,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_module_permission("properties")),
):
    """读取该房源对应的小区详情（admin 后台用）。"""
    from ...models.community import CommunityInfo
    p = (await db.execute(select(Property).where(Property.id == property_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="房源不存在")
    if not p.community_name:
        return {"community_name": "", "info": None}

    c = (await db.execute(
        select(CommunityInfo).where(CommunityInfo.name == p.community_name)
    )).scalar_one_or_none()
    if not c:
        return {"community_name": p.community_name, "info": None}

    return {
        "community_name": p.community_name,
        "info": {
            "id": c.id, "name": c.name, "district": c.district,
            "address_full": c.address_full, "avg_price": c.avg_price,
            "price_update_at": str(c.price_update_at) if c.price_update_at else None,
            "build_year_start": c.build_year_start, "build_year_end": c.build_year_end,
            "property_type": c.property_type, "total_buildings": c.total_buildings,
            "total_units": c.total_units, "developer": c.developer,
            "plot_ratio": c.plot_ratio, "green_rate": c.green_rate,
            "property_company": c.property_company, "property_fee": c.property_fee,
            "huxing_summary": c.huxing_summary,
            "recent_deal_count_30d": c.recent_deal_count_30d,
            "recent_avg_price_30d": c.recent_avg_price_30d,
            "on_sale_count": c.on_sale_count, "rent_count": c.rent_count,
            "description": c.description, "beike_url": c.beike_url,
            "last_crawled_at": str(c.last_crawled_at) if c.last_crawled_at else None,
            "source": c.source,
        },
    }

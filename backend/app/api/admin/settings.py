"""Admin system settings routes."""
import io
import os
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ...core.config import settings as app_settings
from ...core.database import get_session
from ...core.security import get_admin_user, check_write_permission
from ...models.system_setting import SystemSetting
from ...models.property import Property

router = APIRouter()

CITIES = [
    {"city_id": 310000, "city_name": "上海", "is_active": True},
    {"city_id": 330200, "city_name": "宁波", "is_active": True},
    {"city_id": 330100, "city_name": "杭州", "is_active": True},
]

ARCHIVE_DIR = Path(app_settings.IMAGE_STORAGE_PATH) / "archives"

# 全字段中文表头（与 properties /export 一致）
EXPORT_FIELDS = [
    ("id", "ID"), ("title", "标的名称"), ("auction_platform", "拍卖平台"),
    ("source_url", "原始链接"), ("province_city", "省市"), ("district", "区"),
    ("sub_district", "板块"), ("ring_road", "环线"), ("address", "地址"),
    ("community_name", "小区名"), ("property_type", "物业类型"), ("area", "面积(m2)"),
    ("layout", "户型"), ("floor_info", "楼层"), ("total_floors", "总楼层"),
    ("has_elevator", "有无电梯"), ("orientation", "朝向"), ("decoration", "装修情况"),
    ("build_year", "建筑年代"), ("starting_price", "起拍价(元)"),
    ("starting_unit_price", "起拍单价(元/m2)"), ("appraisal_price", "法院评估价(元)"),
    ("court_discount_rate", "法院折扣率"), ("deposit", "保证金(元)"),
    ("increment_amount", "加价幅度(元)"), ("market_deal_price", "市场成交价(元)"),
    ("market_deal_unit_price", "市场成交单价"), ("market_discount_rate", "市场折扣率"),
    ("listing_min_price", "挂牌最低价(元)"), ("latest_deal_unit_price", "最新成交单价"),
    ("latest_total_price", "最新总价(元)"), ("bargain_potential", "捡漏空间(元)"),
    ("beike_latest_deal_unit_price", "贝壳最新成交单价"),
    ("beike_latest_deal_total_price", "贝壳最新成交总价(元)"),
    ("beike_latest_deal_time", "贝壳最新成交时间"),
    ("auction_round", "拍卖轮次"), ("auction_status", "拍卖状态"),
    ("auction_start_time", "开拍时间"), ("auction_end_time", "结束时间"),
    ("court_name", "拍卖法院"), ("case_number", "案号"),
    ("view_count", "围观人数"), ("participant_count", "参拍人数"),
    ("loan_support", "支持贷款"), ("has_attachments", "有无附件"),
    ("publish_date", "发布时间"), ("created_at", "入库时间"), ("updated_at", "更新时间"),
]


def _generate_csv_content(rows) -> str:
    output = io.StringIO()
    output.write(",".join(h for _, h in EXPORT_FIELDS) + "\n")
    for p in rows:
        vals = []
        for key, _ in EXPORT_FIELDS:
            v = getattr(p, key, "")
            if v is None:
                v = ""
            elif isinstance(v, bool):
                v = "是" if v else "否"
            v = str(v).replace('"', '""')
            vals.append(f'"{v}"')
        output.write(",".join(vals) + "\n")
    content = output.getvalue()
    output.close()
    return content


@router.get("")
async def get_settings(
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(get_admin_user),
):
    rows = (await db.execute(select(SystemSetting))).scalars().all()
    return {r.key: r.value for r in rows}


@router.put("")
async def update_settings(
    body: dict,
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
):
    for key, value in body.items():
        result = await db.execute(
            select(SystemSetting).where(SystemSetting.key == key)
        )
        setting = result.scalar_one_or_none()
        if setting:
            setting.value = str(value) if value is not None else ""
        else:
            db.add(SystemSetting(key=key, value=str(value) if value is not None else ""))
    await db.commit()
    return {"message": "配置已更新"}


@router.get("/cities")
async def list_cities(admin: dict = Depends(get_admin_user)):
    return CITIES


@router.post("/cities")
async def add_city(
    body: dict,
    admin: dict = Depends(check_write_permission()),
):
    city_id = body.get("city_id")
    city_name = body.get("city_name", "")
    if not city_id or not city_name:
        raise HTTPException(status_code=400, detail="city_id 和 city_name 必填")
    city = {"city_id": city_id, "city_name": city_name, "is_active": True}
    CITIES.append(city)
    return {"message": "城市已添加", "city": city}


# ============ 数据归档 ============


def _generate_xlsx_bytes(rows) -> bytes:
    """生成 xlsx 内容（专业格式：表头加粗+底色，列宽自适应，冻结表头）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date

    wb = Workbook()
    ws = wb.active
    ws.title = "房源归档"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_, label) in enumerate(EXPORT_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[1].height = 28

    body_align = Alignment(horizontal="left", vertical="center")
    for r_idx, p in enumerate(rows, start=2):
        for c_idx, (key, _) in enumerate(EXPORT_FIELDS, start=1):
            v = getattr(p, key, "")
            if v is None:
                cell_val = ""
            elif isinstance(v, bool):
                cell_val = "是" if v else "否"
            elif isinstance(v, (datetime, date, int, float)):
                cell_val = v
            else:
                cell_val = str(v)
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.alignment = body_align
            cell.border = border
            if isinstance(v, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(v, date):
                cell.number_format = "yyyy-mm-dd"

    for col_idx, (key, _) in enumerate(EXPORT_FIELDS, start=1):
        col_letter = get_column_letter(col_idx)
        if key in ("title", "address", "community_name", "court_name"):
            width = 36
        elif key in ("source_url", "announcement_url"):
            width = 40
        elif "price" in key or "rate" in key or "amount" in key:
            width = 14
        elif "_at" in key or "_date" in key or "_time" in key:
            width = 18
        else:
            width = 12
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.post("/archive/export")
async def manual_archive_export(
    db: AsyncSession = Depends(get_session),
    admin: dict = Depends(check_write_permission()),
    format: str = "xlsx",
):
    """手动导出当前全量房源数据为归档文件（xlsx 默认，可选 csv）。"""
    rows = (await db.execute(select(Property).order_by(Property.id))).scalars().all()

    now = datetime.now()
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    if format == "csv":
        content = _generate_csv_content(rows)
        filename = f"房源归档_{now.strftime('%Y%m%d_%H%M%S')}_{len(rows)}条.csv"
        filepath = ARCHIVE_DIR / filename
        filepath.write_text(content, encoding="utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # 默认 xlsx
    xlsx_bytes = _generate_xlsx_bytes(rows)
    filename = f"房源归档_{now.strftime('%Y%m%d_%H%M%S')}_{len(rows)}条.xlsx"
    filepath = ARCHIVE_DIR / filename
    filepath.write_bytes(xlsx_bytes)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/archive/list")
async def list_archives(admin: dict = Depends(get_admin_user)):
    """列出已有的归档文件（CSV + XLSX）。"""
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    files = []
    seen = set()
    for pattern in ("房源归档_*.xlsx", "房源归档_*.csv"):
        for f in ARCHIVE_DIR.glob(pattern):
            if f.name in seen:
                continue
            seen.add(f.name)
            stat = f.stat()
            files.append({
                "filename": f.name,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return files


@router.get("/archive/download/{filename}")
async def download_archive(
    filename: str,
    admin: dict = Depends(get_admin_user),
):
    """下载指定归档文件。"""
    filepath = ARCHIVE_DIR / filename
    if not filepath.exists() or not filepath.name.startswith("房源归档_"):
        raise HTTPException(status_code=404, detail="归档文件不存在")
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if filename.endswith(".xlsx") else "text/csv"
    )
    return FileResponse(filepath, media_type=media_type, filename=filename)


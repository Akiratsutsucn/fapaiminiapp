# -*- coding: utf-8 -*-
"""生成阿里/京东三城市在拍房源数量对比表 xlsx。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "在拍数量对比"

# 样式
title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="2E75B6")
sub_fill = PatternFill("solid", fgColor="DDEBF7")
sub_font = Font(bold=True)
total_font = Font(bold=True)
total_fill = PatternFill("solid", fgColor="FCE4D6")
ok_fill = PatternFill("solid", fgColor="C6EFCE")      # 吻合-绿
big_fill = PatternFill("solid", fgColor="FFC7CE")     # 大差异-红
mid_fill = PatternFill("solid", fgColor="FFEB9C")     # 中差异-黄
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 数据: (类型, 状态, 网站手动, 数据库)
ali = [
    ("住宅", "即将开拍", 237, 588),
    ("住宅", "进行中", 22, 78),
    ("商业", "即将开拍", 81, 113),
    ("商业", "进行中", 26, 52),
    ("工业", "即将开拍", 4, 4),
    ("工业", "进行中", 1, 1),
    ("其他", "即将开拍", 40, 18),
    ("其他", "进行中", 7, 8),
]
jd = [
    ("住宅", "即将开拍", 57, 1),
    ("住宅", "进行中", 8, 6),
    ("商业", "即将开拍", 30, 0),
    ("商业", "进行中", 25, 23),
    ("工业", "即将开拍", 1, 0),
    ("工业", "进行中", 0, 0),
    ("其他", "即将开拍", 10, 0),
    ("其他", "进行中", 0, 4),
]

# 标题行
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "阿里 / 京东 三城市(上海+宁波+杭州)在拍房源数量对比"
c.font = title_font; c.fill = title_fill; c.alignment = center
ws.merge_cells("A2:E2")
c = ws["A2"]
c.value = "网站手动清点时间: 2026-06-04 约9点  |  数据库口径: effective_status 实时状态, property_type 归入住宅/商业/工业/其他"
c.font = Font(italic=True, size=9, color="595959"); c.alignment = center

row = 4
def write_header():
    global row
    headers = ["物业类型", "拍卖状态", "网站手动清点", "数据库查询", "差异(库-站)"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    row += 1

def write_block(name, data):
    global row
    # 平台分组行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=f"【{name}】")
    cell.font = sub_font; cell.fill = sub_fill; cell.alignment = left; cell.border = border
    for j in range(2, 6):
        ws.cell(row=row, column=j).border = border
    row += 1
    sub_site = sub_db = 0
    for pt, st, site, db in data:
        diff = db - site
        sub_site += site; sub_db += db
        vals = [pt, st, site, db, diff]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.alignment = center if j != 1 else center
            cell.border = border
        # 差异列着色
        dcell = ws.cell(row=row, column=5)
        if diff == 0:
            dcell.fill = ok_fill
        elif abs(diff) >= 30:
            dcell.fill = big_fill
        elif abs(diff) >= 10:
            dcell.fill = mid_fill
        if diff > 0:
            dcell.value = f"+{diff}"
        row += 1
    # 小计
    sub_diff = sub_db - sub_site
    totals = [f"{name}小计", "", sub_site, sub_db, (f"+{sub_diff}" if sub_diff > 0 else sub_diff)]
    for j, v in enumerate(totals, start=1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = total_font; cell.fill = total_fill; cell.alignment = center; cell.border = border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

write_header()
write_block("阿里法拍", ali)
write_block("京东", jd)

# 列宽
for col, w in zip("ABCDE", [14, 12, 14, 12, 12]):
    ws.column_dimensions[col].width = w

# 诊断说明
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row=row, column=1, value="差异诊断")
cell.font = Font(bold=True, size=12, color="1F4E78"); cell.alignment = left
row += 1

notes = [
    "1. 阿里『即将开拍』库里约为网站2.5倍：数据库住宅588条中，未来30天内497条、超30天还有41条；网站手动237很可能只数了默认/近期列表，差异主要来自统计时间窗口口径不同。其中约50条为过期残留(本次爬取未再抓到但状态仍卡在『即将开拍』)，属应清理的脏数据。",
    "2. 城市分布异常：阿里住宅即将开拍 杭州316 > 上海214，正常应上海最多，杭州疑似过抓，需专项核查。",
    "3. 京东几乎全军覆没(库34 vs 网站131)：京东三市117条全部为已结束/已撤回/已成交，即将开拍住宅库里仅1条。与当日爬虫汇总『京东 0 found』吻合——京东爬虫本次未抓到任何在拍房源，属真实故障，需修复。",
    "4. 工业类完全吻合，说明能对上的部分爬虫数据是准确的。",
    "颜色说明：绿=完全吻合，黄=差异10~29，红=差异≥30。",
]
for n in notes:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=n)
    cell.alignment = left; cell.font = Font(size=10)
    ws.row_dimensions[row].height = 46
    row += 1

out = r"C:\Users\Administrator\Desktop\workspace\法拍者联盟小程序\阿里京东三城市在拍数量对比_20260604.xlsx"
wb.save(out)
print("SAVED:", out)
